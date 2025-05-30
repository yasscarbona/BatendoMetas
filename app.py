#Importando os módulos necessários para o Flask funcionar
from flask import Flask, render_template,request,redirect
#Importando as funções da biblioteca openpyxl para criar e
#manipular um arquivo excel
from openpyxl import Workbook, load_workbook
#Biblioteca para verificar a existência de um arq. excel
import os
#Criar de fato a nossa aplicação
app = Flask(__name__)
#Definindo o nome da planilha excel
ARQUIVO = 'turma.xlsx'

if not os.path.exists(ARQUIVO):
    wb = Workbook()#CRIANDO UM ARQ. EXCEL
    ws = wb.active #Selecionando um plano. ativa do projeto
#Criando um cabeçalho para a planilha
    ws.append(["Nome","Media","Situacao"])
#salvando o arquivo
    wb.save(ARQUIVO)

#22/05
#Rota principal do site (formulário cadastro de vendas)
@app.route('/')
def index():
    return render_template('index.html')
#função que deve ser executada, ao ser requisitado a rota '/'
#abrir a página principal, neste caso é o index.html

#Rota que processa os dados do formulário e salva no Excel
@app.route('/salvar',methods=["POST"])
def salvar():
#CAPTURA OS DADOS DE CADA UMA DAS CAIXAS DO FORMULÁRIO E ATRIBUI PARA
#AS VARIÁVEIS
    nome = request.form['nome']
#CAPTURA A INFO. DA CX. FAÇA O FORMULÁRIO. CONVERTE PARA FLOAT E ATRIBUI À
#VARIÁVEL
    nota01 = float(request.form['nota01'])
    nota02 = float(request.form['nota02'])
#ABRINDO O ARQUIVO EXCEL
    wb = load_workbook(ARQUIVO)
#SELECIONANDO A PLANILHA ATIVA - 1ª ABA Por padrão
    ws = wb.active
# Tem de ter o calculo da media e o if para a situação
    media = round((nota01 + nota02) /2,2)
    situacao = "Aprovado" if media >= 6 else "Reprovado!"
    ws.append([nome,nota01, nota02, media,situacao])

#Salvando o arquivo excel
    wb.save(ARQUIVO)

#Redirecionando a rota para /analisar (onde abrirá uma nova página)
#Passando por parâmetro o nome do funcionário.
    return redirect('/analisar?nome='+nome)
#Rota para a tela resultado... analisando se o funcionário bateu a meta
@app.route('/analisar')
def analisar():
#Pegar o parâmetro do nome da func. enviado como parâmetro para url
    nome_param = request.args.get('nome')

    wb = load_workbook(ARQUIVO)
    ws = wb.active

#Loop for: percorre todas as linhas do plano. a partir da 2ª linha
#(Pois a 1ª linha é cabeçalho) valores_only retorna apenas os valores
#da linha em
    for linha in ws.iter_rows(min_row=2, values_only=True):
        nome,nota01,nota02, media,situacao = linha
    # a linha sempre recebe três valores, ex:
    # linha = ('Ana', 45, 50) na linha de código acima, estamos
    # atribuindo cada elemento da linha a uma variável na sequencia
    # as variáveis ​​nome, vendas e meta recebem Ana, 45, 50
    # respectivamente. Nome técnico desse processo é desempacotamento

    #Verifique se o nome atual da linha é o mesmo enviado na URL
        if nome == nome_param:
            media = round((nota01 + nota02)/2,2)
            situacao = "Aprovado" if media >= 6 else "Reprovado!"
        #Exibir a tela resultante com as informações dos cálculos
            return render_template('resultado.html',
            nome=nome,media=media,situacao=situacao)
    return "Aluno não encontrado"

#Rota que mostra a pág. do histórico de todos os funcionários cadastrados
@app.route('/historico')
def historico():
    wb = load_workbook(ARQUIVO)
    ws = wb.active
#Converte os dados da planilha ( a partir da 2ª linha ) em uma tupla
    dados = list(ws.iter_rows(min_row = 2,values_only=True))
    return render_template('historico.html', dados = dados)
    #Iniciando o Flask no modo desenvolvedor Debug
if __name__ == '__main__':
        app.run(debug=True)